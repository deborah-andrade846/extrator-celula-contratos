# -*- coding: utf-8 -*-
"""
Extrator de Relatórios - Apoena
Versão Unificada: Fiscal (OCR + NAI) + Hotel + Exames + Refeições
"""

import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pytesseract
import cv2
import numpy as np
from typing import List, Tuple, Dict, Optional
from dataclasses import dataclass, asdict
import logging
from contextlib import contextmanager
import gc
import unicodedata

# ==================== CONFIGURAÇÕES GLOBAIS ====================
COLUNAS_CONFIG = {
    "hotel": ["Arquivo", "Data", "Informação adicional", "Qtde", "Unidade", "Total"],
    "exames": ["Arquivo", "Exame", "Valor"],
    "refeicoes": ["Arquivo", "Data", "Total"],
    "refeicoes_empresas": ["Arquivo", "Empresa", "Total"],
    "fiscal": [
        "Arquivo", "Tipo NAI", "Data", "Nº NF", "Chave da NF-e", "Fornecedor", 
        "UF", "NCM", "Descrição", "CFOP", "Valor NF", "BC ICMS", "ICMS", 
        "% Interna", "ICMS Origem", "VR DIFAL", "OBS"
    ]
}

# ==================== LOGGING ====================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== DATACLASS PARA ESTATÍSTICAS ====================
@dataclass
class EstatisticasProcessamento:
    arquivo: str
    total_linhas_encontradas: int = 0
    sucesso: int = 0
    falhas: int = 0

    def taxa_sucesso(self) -> float:
        if self.total_linhas_encontradas == 0:
            return 0.0
        return (self.sucesso / self.total_linhas_encontradas) * 100

# ==================== GERENCIADOR DE MEMÓRIA ====================
@contextmanager
def gerenciar_memoria(limite_mb: int = 500):
    try:
        yield
    finally:
        gc.collect()

# ==================== FUNÇÕES AUXILIARES ====================
def converter_para_numero(valor_str):
    if not valor_str or valor_str == "N/D":
        return None
    try:
        return float(valor_str.replace('.', '').replace(',', '.'))
    except:
        return valor_str

# ==================== OCR COM VISÃO COMPUTACIONAL ====================
# Confiança mínima do OSD para aceitar uma rotação. O Tesseract às vezes sugere
# "girar 180°" para páginas já corretas com confiança ínfima (ex.: 0.01); confiar
# nesse palpite vira uma página boa de cabeça para baixo e destrói o OCR.
OSD_CONF_MINIMA = 2.0

def corrigir_rotacao(img_cinza):
    """Detecta orientação via OSD em resolução reduzida (rápido) e corrige a imagem original.

    Só aplica a rotação quando o OSD tem confiança suficiente; caso contrário mantém
    a imagem original (que, na prática, quase sempre já está na orientação correta).
    """
    try:
        img_baixa = cv2.resize(img_cinza, None, fx=0.5, fy=0.5)
        osd = pytesseract.image_to_osd(img_baixa, output_type=pytesseract.Output.DICT)
        rotate = osd.get('rotate', 0)
        conf = float(osd.get('orientation_conf', 0) or 0)
        if rotate and conf < OSD_CONF_MINIMA:
            return img_cinza  # palpite pouco confiável: não arrisca girar
        if rotate == 90:
            return cv2.rotate(img_cinza, cv2.ROTATE_90_CLOCKWISE)
        elif rotate == 180:
            return cv2.rotate(img_cinza, cv2.ROTATE_180)
        elif rotate == 270:
            return cv2.rotate(img_cinza, cv2.ROTATE_90_COUNTERCLOCKWISE)
    except Exception:
        pass
    return img_cinza

_ROTACOES_CV = {
    90: cv2.ROTATE_90_CLOCKWISE,
    180: cv2.ROTATE_180,
    270: cv2.ROTATE_90_COUNTERCLOCKWISE,
}

def _preparar_imagem_ocr(pagina, rotacao=None):
    """Rasteriza, corrige a rotação e binariza uma página para OCR de alta precisão.

    Devolve a imagem binarizada (numpy) para que os consumidores possam tanto extrair
    o texto puro quanto os dados posicionais (image_to_data) a partir da mesma base.

    Quando ``rotacao`` (0/90/180/270) é informada, aplica exatamente essa rotação; caso
    contrário decide página a página via OSD (``corrigir_rotacao``).
    """
    img_pil = pagina.to_image(resolution=300).original
    img_cv = np.array(img_pil)
    if len(img_cv.shape) == 3:
        img_cv = cv2.cvtColor(img_cv, cv2.COLOR_RGB2BGR)
    img_cinza = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    if rotacao is None:
        img_cinza = corrigir_rotacao(img_cinza)
    elif rotacao in _ROTACOES_CV:
        img_cinza = cv2.rotate(img_cinza, _ROTACOES_CV[rotacao])
    img_suave = cv2.GaussianBlur(img_cinza, (3, 3), 0)
    _, img_bin = cv2.threshold(img_suave, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    return img_bin

def _detectar_rotacao_documento(pdf, amostras=5):
    """Descobre a rotação predominante do documento com uma votação de OSD.

    O OSD por página às vezes devolve confiança baixa em páginas boas (deixando-as sem
    girar e ilegíveis). Como todas as páginas do relatório têm a mesma orientação,
    amostramos as primeiras páginas, somamos a confiança de cada rotação sugerida e
    aplicamos o vencedor a todas — bem mais robusto que decidir página a página.
    """
    from collections import defaultdict
    votos = defaultdict(float)
    paginas = pdf.pages[:amostras] if len(pdf.pages) >= amostras else pdf.pages
    for pagina in paginas:
        try:
            img = cv2.cvtColor(np.array(pagina.to_image(resolution=150).original), cv2.COLOR_RGB2GRAY)
            osd = pytesseract.image_to_osd(img, output_type=pytesseract.Output.DICT)
            votos[int(osd.get('rotate', 0))] += float(osd.get('orientation_conf', 0) or 0)
        except Exception:
            pass
    return max(votos, key=votos.get) if votos else 0

def ocr_pagina(pagina) -> str:
    """Aplica Visão Computacional + OCR de alta precisão numa única página do pdfplumber."""
    return pytesseract.image_to_string(_preparar_imagem_ocr(pagina), lang='por+eng', config='--psm 6 --oem 3')

def ler_texto_com_ocr(arquivo_pdf):
    texto_completo = ""
    barra_progresso = st.progress(0, text="A aplicar Visão Computacional e OCR...")
    with pdfplumber.open(arquivo_pdf) as pdf:
        total_paginas = len(pdf.pages)
        for i, pagina in enumerate(pdf.pages):
            texto_completo += ocr_pagina(pagina) + "\n"
            barra_progresso.progress((i + 1) / total_paginas, text=f"OCR: Página {i+1} de {total_paginas}")
    barra_progresso.empty()
    return texto_completo

# ==================== PARSER FISCAL ====================
class FiscalParser:
    PADROES = {
        '89701': re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(\d{44})\s+(.*?)\s+([A-Z]{2})\s+(.*?)\s+(\d{4})\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$'),
        '92284': re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(\d{44})\s+([A-Z]{2})\s+(\d{8})\s+(\d{4})\s+(.*?)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*(.*?)\s*$'),
        'MISTO': re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(\d{44})\s+(.*?)\s+([A-Z]{2})\s+(\d{8})\s+(.*?)\s+(\d{4})\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$')
    }

    @classmethod
    def parsear_linha(cls, linha: str, arquivo: str) -> Optional[Dict]:
        if not re.match(r'^\d{2}/\d{2}/\d{4}\s+\d+\s+\d{44}', linha):
            return None
        for tipo, padrao in cls.PADROES.items():
            m = padrao.match(linha)
            if m:
                return cls._montar_registro(m, tipo, arquivo)
        return None

    @staticmethod
    def _montar_registro(m, tipo, arquivo):
        if tipo == 'MISTO':
            return {
                "Arquivo": arquivo, "Tipo NAI": "Misto", "Data": m.group(1), "Nº NF": m.group(2),
                "Chave da NF-e": m.group(3), "Fornecedor": m.group(4).strip(), "UF": m.group(5),
                "NCM": m.group(6), "Descrição": m.group(7).strip(), "CFOP": m.group(8),
                "Valor NF": converter_para_numero(m.group(9)), "BC ICMS": converter_para_numero(m.group(10)),
                "ICMS": converter_para_numero(m.group(11)), "% Interna": converter_para_numero(m.group(12)),
                "ICMS Origem": converter_para_numero(m.group(13)), "VR DIFAL": converter_para_numero(m.group(14)), "OBS": ""
            }
        elif tipo == '89701':
            return {
                "Arquivo": arquivo, "Tipo NAI": "89701", "Data": m.group(1), "Nº NF": m.group(2),
                "Chave da NF-e": m.group(3), "Fornecedor": m.group(4).strip(), "UF": m.group(5),
                "NCM": "", "Descrição": m.group(6).strip(), "CFOP": m.group(7),
                "Valor NF": converter_para_numero(m.group(8)), "BC ICMS": None, "ICMS": None,
                "% Interna": None, "ICMS Origem": converter_para_numero(m.group(9)),
                "VR DIFAL": converter_para_numero(m.group(10)), "OBS": ""
            }
        else:  # 92284
            return {
                "Arquivo": arquivo, "Tipo NAI": "92284", "Data": m.group(1), "Nº NF": m.group(2),
                "Chave da NF-e": m.group(3), "Fornecedor": "", "UF": m.group(4),
                "NCM": m.group(5), "Descrição": m.group(7).strip(), "CFOP": m.group(6),
                "Valor NF": converter_para_numero(m.group(8)), "BC ICMS": converter_para_numero(m.group(9)),
                "ICMS": converter_para_numero(m.group(10)), "% Interna": converter_para_numero(m.group(11)),
                "ICMS Origem": None, "VR DIFAL": converter_para_numero(m.group(12)), "OBS": m.group(13).strip()
            }

# ==================== EXTRAÇÃO FISCAL COMPLETA ====================
def extrair_fiscal(arquivo_pdf, usar_ocr: bool):
    dados_locais, rejeitadas = [], []
    stats = EstatisticasProcessamento(arquivo=arquivo_pdf.name)

    if usar_ocr:
        texto = ler_texto_com_ocr(arquivo_pdf)
    else:
        texto = ""
        with pdfplumber.open(arquivo_pdf) as pdf:
            for pagina in pdf.pages:
                txt = pagina.extract_text(layout=True)
                if txt:
                    texto += txt + "\n"

    linhas_brutas = texto.split('\n')
    linhas_limpas = []
    for linha in linhas_brutas:
        linha = re.sub(r'(?<!^)(?=\d{2}/\d{2}/\d{4}\s+\d+\s+\d{44})', '\n', linha)
        for sub in linha.split('\n'):
            sub = re.sub(r'[|—\[\]]', '', sub)
            sub = re.sub(r'\b([A-Z]{2})\.', r'\1', sub)
            linhas_limpas.append(sub.strip())

    for linha in linhas_limpas:
        if not linha or not re.match(r'^\d{2}/\d{2}/\d{4}\s+\d+\s+\d{44}', linha):
            continue
        stats.total_linhas_encontradas += 1
        registro = FiscalParser.parsear_linha(linha, arquivo_pdf.name)
        if registro:
            dados_locais.append(registro)
            stats.sucesso += 1
        else:
            stats.falhas += 1
            rejeitadas.append({"Arquivo": arquivo_pdf.name, "Linha Completa Não Processada": linha})

    return dados_locais, rejeitadas, stats

# ==================== EXTRAÇÃO HOTEL ====================
# Token monetário no formato brasileiro (15,00 / 1.234,56) ou lido pelo OCR com
# ponto no lugar da vírgula (15.00 / 1.234.56). Sempre com 2 casas decimais, o que
# evita capturar códigos ou números soltos da descrição.
_TOKEN_VALOR_HOTEL = re.compile(r'^\d{1,3}(?:\.\d{3})*[.,]\d{2}$')

def _normalizar_valor_hotel(valor: str) -> str:
    """Padroniza o separador decimal para vírgula (formato brasileiro)."""
    if valor.count('.') and valor.count(','):
        # Ex.: 1.234,56 -> mantém
        return valor
    # Ex.: 15.00 (OCR leu ponto) -> 15,00
    return valor.replace('.', ',') if valor.count('.') == 1 and ',' not in valor else valor

def limpar_linha_hotel(linha, nome_hospede):
    linha = linha.strip()
    # 1. A linha precisa começar por uma data. O dia/mês podem vir com 1 dígito
    #    porque o OCR às vezes perde um algarismo (ex.: lê "1/06/26" em vez de
    #    "11/06/26"); sem isso a linha seria descartada.
    match = re.match(r'^(\d{1,2}/\d{1,2}/\d{2,4})\b', linha)
    if not match:
        return None
    data = match.group(1)
    resto = linha[match.end():].strip()
    # 2. Remove um horário opcional logo após a data. Aceita HH:MM / HH:MM:SS e
    #    também formas em que o OCR perdeu o separador ou trocou por ponto
    #    (ex.: "1412", "0915", "10.11").
    resto = re.sub(r'^\d{1,2}[:.]?\d{2}(?::\d{2})?\s*', '', resto)
    partes = resto.split()
    # 3. Coleta os tokens numéricos finais (Qtde Unidade Bruto Desc. Taxas Total)
    numericos = []
    i = len(partes) - 1
    while i >= 0 and _TOKEN_VALOR_HOTEL.match(partes[i]):
        numericos.insert(0, partes[i])
        i -= 1
    # Precisa ao menos de Qtde, Unidade e Total para ser uma linha de consumo/diária
    if len(numericos) < 3:
        return None
    # 4. Monta a informação adicional (tudo antes dos números) e captura a Comanda
    info_completa = " ".join(partes[:i + 1]).replace("|", "-").strip()
    m_comanda = re.search(r'Comanda\s+([A-Za-z]*\d+)', info_completa, re.IGNORECASE)
    comanda = m_comanda.group(1).upper() if m_comanda else None
    info = re.split(r'\s*-\s*Comanda', info_completa, flags=re.IGNORECASE)[0].strip()
    return {
        "Arquivo": nome_hospede,
        "Data": data,
        "Informação adicional": info,
        "Qtde": _normalizar_valor_hotel(numericos[0]),
        "Unidade": _normalizar_valor_hotel(numericos[1]),
        "Total": _normalizar_valor_hotel(numericos[-1]),
        "_comanda": comanda
    }

def _corrigir_datas_por_comanda(dados):
    """Corrige datas quebradas pelo OCR usando a Comanda como âncora.

    Todos os itens de uma mesma Comanda pertencem ao mesmo lançamento (mesma data).
    Quando o OCR perde um dígito do dia (ex.: "1/06/26" em vez de "11/06/26"),
    adotamos a data predominante da Comanda, preferindo o formato com dia/mês de
    2 dígitos em caso de empate.
    """
    from collections import defaultdict
    grupos = defaultdict(list)
    for d in dados:
        if d.get("_comanda"):
            grupos[d["_comanda"]].append(d)
    canonica = re.compile(r'^\d{2}/\d{2}/\d{2,4}$')
    for linhas in grupos.values():
        datas = [l["Data"] for l in linhas]
        melhor = max(datas, key=lambda dt: (datas.count(dt), 1 if canonica.match(dt) else 0))
        for l in linhas:
            l["Data"] = melhor
    for d in dados:
        d.pop("_comanda", None)
    return dados

def extrair_hotel(arquivo_pdf, usar_ocr=False):
    dados = []
    nome_hospede = "NÃO_IDENTIFICADO"
    with pdfplumber.open(arquivo_pdf) as pdf:
        # 1. Tenta texto direto (rápido)
        texto_completo = "\n".join([pagina.extract_text() or "" for pagina in pdf.pages])

        # 2. Aplica OCR de alta precisão se solicitado ou se o texto direto for insuficiente
        #    (PDF escaneado / sem camada de texto legível)
        texto_direto_suficiente = "Hóspede principal:" in texto_completo and texto_completo.strip() != ""
        if usar_ocr or not texto_direto_suficiente:
            st.info(f"Aplicando OCR de alta precisão em {arquivo_pdf.name}...")
            total_paginas = len(pdf.pages)
            barra = st.progress(0, text="OCR nas diárias...")
            texto_completo = ""
            for i, pagina in enumerate(pdf.pages):
                texto_completo += ocr_pagina(pagina) + "\n"
                barra.progress((i + 1) / total_paginas, text=f"OCR: Página {i+1} de {total_paginas}")
            barra.empty()

        linhas = texto_completo.split('\n')
        for linha in linhas:
            if "Hóspede principal:" in linha:
                try:
                    nome_cru = linha.split("Hóspede principal:")[1].split("|")[0].strip()
                    nome_hospede = nome_cru.split()[0]
                except:
                    pass
                continue
            if any(palavra in linha for palavra in ["PLAZA HOTEL", "Apartamento:", "Fechado", "Pagamentos", "Tarifário:"]):
                continue
            linha_extraida = limpar_linha_hotel(linha, nome_hospede)
            if linha_extraida:
                dados.append(linha_extraida)
    return _corrigir_datas_por_comanda(dados)

# ==================== EXTRAÇÃO EXAMES ====================
def extrair_exames(arquivo_pdf):
    dados = []
    with pdfplumber.open(arquivo_pdf) as pdf:
        texto_completo = "\n".join([pagina.extract_text() or "" for pagina in pdf.pages])
        linhas = texto_completo.split('\n')
        for linha in linhas:
            if "R$" in linha:
                partes = linha.split("R$")
                if len(partes) >= 2:
                    nome_exame = partes[0].replace('"', '').replace(',', '').strip()
                    valor_raw = partes[1].replace('"', '').strip()
                    try:
                        valor_limpo = valor_raw.replace('.', '').replace(',', '.')
                        valor_num = float(valor_limpo)
                        valor_formatado = f"R$ {valor_num:,.2f}".replace('.', ',')
                    except:
                        valor_formatado = f"R$ {valor_raw}"
                    
                    if nome_exame:
                        dados.append({
                            "Arquivo": arquivo_pdf.name,
                            "Exame": nome_exame,
                            "Valor": valor_formatado
                        })
    return dados

# ==================== EXTRAÇÃO REFEIÇÕES (ROBUSTA COM OCR) ====================
def extrair_refeicoes(arquivo_pdf, usar_ocr=False):
    dados = []
    data_refeicao = "DATA_NAO_ENCONTRADA"
    total_refeicoes = "TOTAL_NAO_ENCONTRADO"
    
    # 1. Tenta texto direto — lê só a primeira página (data/período) e a última (Total Geral)
    texto_primeira = ""
    texto_ultima = ""
    texto_direto_suficiente = False
    with pdfplumber.open(arquivo_pdf) as pdf:
        txt_primeira = pdf.pages[0].extract_text()
        if txt_primeira:
            texto_primeira = txt_primeira
        txt_ultima = pdf.pages[-1].extract_text()
        if txt_ultima:
            texto_ultima = txt_ultima

    texto_completo = texto_primeira + "\n" + texto_ultima

    if re.search(r'per[ií]odo', texto_primeira, re.IGNORECASE) and re.search(r'total\s*geral', texto_ultima, re.IGNORECASE):
        texto_direto_suficiente = True

    # 2. Se necessário, aplica OCR apenas nas páginas relevantes (primeira e última)
    if usar_ocr or not texto_direto_suficiente:
        st.info(f"Aplicando OCR em {arquivo_pdf.name}...")
        with pdfplumber.open(arquivo_pdf) as pdf:
            indices = list({0, len(pdf.pages) - 1})  # primeira e última (evita duplicar se for 1 página)
            texto_completo = ""
            barra = st.progress(0, text="OCR nas páginas relevantes...")
            for j, idx in enumerate(indices):
                texto_completo += ocr_pagina(pdf.pages[idx]) + "\n"
                barra.progress((j + 1) / len(indices), text=f"OCR: página {idx + 1}")
            barra.empty()
    
    # ---- DEBUG (opcional) ----
    # st.text_area("Texto bruto do OCR", texto_completo[:3000], height=200)
    
    # Limpeza de ruídos comuns de OCR
    texto_limpo = texto_completo.replace('|', '').replace('—', '-').replace('_', '')
    linhas = texto_limpo.split('\n')
    
    # Padrões flexíveis
    data_pattern = r'(\d{2}[/.-]\d{2}[/.-]\d{2,4})'
    # Aceita tanto números com decimais (1.234,56) quanto inteiros simples (436)
    valor_pattern = r'(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})|\d+)'
    
    for i, linha in enumerate(linhas):
        # Procura "Período" (com ou sem acento)
        if re.search(r'per[ií]odo', linha, re.IGNORECASE):
            match_data = re.search(data_pattern, linha)
            if match_data:
                data_refeicao = match_data.group(1).replace('-', '/')
            elif i+1 < len(linhas):
                match_data = re.search(data_pattern, linhas[i+1])
                if match_data:
                    data_refeicao = match_data.group(1).replace('-', '/')
        
        # Procura "Total Geral"
        if re.search(r'total\s*geral', linha, re.IGNORECASE):
            match_valor = re.search(valor_pattern, linha)
            if match_valor:
                total_refeicoes = match_valor.group(1).replace('.', ',').lstrip('0')  # ajusta formato
            else:
                for offset in range(1, 3):
                    if i+offset < len(linhas):
                        match_valor = re.search(valor_pattern, linhas[i+offset])
                        if match_valor:
                            total_refeicoes = match_valor.group(1).replace('.', ',').lstrip('0')
                            break
    
    # Fallback: se não encontrou data, tenta a primeira data do documento
    if data_refeicao == "DATA_NAO_ENCONTRADA":
        match_data = re.search(data_pattern, texto_limpo)
        if match_data:
            data_refeicao = match_data.group(1).replace('-', '/')
    
    # Se encontrou algo, cria o registro
    if data_refeicao != "DATA_NAO_ENCONTRADA" or total_refeicoes != "TOTAL_NAO_ENCONTRADO":
        dados.append({
            "Arquivo": arquivo_pdf.name,
            "Data": data_refeicao,
            "Total": total_refeicoes
        })
    
    return dados

# ==================== REFEIÇÕES: TOTAL POR EMPRESA ====================
# No Mapa Gerencial de Refeições as linhas são hierárquicas e são identificadas pelo
# recuo (indentação) horizontal: as categorias de vínculo (Matriculado/Visitante/
# Contratado) ficam mais à esquerda, seguidas das EMPRESAS, depois os setores e, por
# fim, as pessoas. Medimos o recuo como fração da largura da página para ficar
# independente da resolução; a faixa abaixo isola justamente o nível das empresas.
_RECUO_EMPRESA_MIN = 0.028
_RECUO_EMPRESA_MAX = 0.043
# A coluna de quantidade fica alinhada à direita da página; só tratamos como total os
# números que aparecem a partir deste recuo.
_RECUO_COLUNA_QTDE_MIN = 0.18
# Nomes de empresa com mais tokens do que isto são, na prática, ruído de OCR.
_MAX_TOKENS_EMPRESA = 8

# Trechos de rodapé/cabeçalho que às vezes escorregam para a faixa das empresas.
_LIXO_KEYWORDS = (
    'copia', 'licenciada', 'n.p.j', 'cnpj', 'foracesso',
    'gerencial de refeicoes', 'periodo:', 'pagina', 'usuario', 'relatorio de mapa',
)

def _sem_acento(texto):
    return ''.join(c for c in unicodedata.normalize('NFKD', texto) if not unicodedata.combining(c))

def _linha_de_rodape(texto):
    """True para linhas de rodapé/cabeçalho (não são empresas)."""
    t = _sem_acento(texto).lower()
    return any(k in t for k in _LIXO_KEYWORDS)

def _token_lixo(token):
    """Token que deve ser removido do fim de um nome de empresa.

    Nomes de empresa vêm em CAIXA ALTA; um token com minúscula só é aceito se for uma
    palavra em Título com pelo menos 4 letras (ex.: 'Total', 'Vigilância'). O restante
    (fragmentos minúsculos, símbolos soltos) é ruído de OCR.
    """
    if re.fullmatch(r'[A-ZÀ-Ý][a-zà-ÿç]{3,}', token):
        return False
    if re.search(r'[a-zà-ÿ]', token):
        return True
    if re.search(r'[^0-9A-ZÀ-ÝÇ&/.\-]', token):
        return True
    return False

def _nome_empresa_plausivel(nome):
    if re.search(r'\b0*\d{3,5}\b', nome):        # centro de custo (00002, 00017, ...)
        return True
    if re.search(r'[A-ZÀ-ÝÇ]{4,}', nome):        # palavra MAIÚSCULA com 4+ letras
        return True
    fortes = re.findall(r'[A-ZÀ-Ý][a-zà-ÿç]{2,}', nome)
    return len(fortes) >= 2                        # ex.: 'Total Vigilância'

def _limpar_nome_empresa(texto):
    """Normaliza o texto de uma linha de empresa e rejeita ruído de OCR (devolve '')."""
    tokens = texto.split()
    while tokens and _token_lixo(tokens[-1]):
        tokens.pop()
    nome = re.sub(r'\s+', ' ', ' '.join(tokens)).strip()
    if not nome:
        return ''
    # Um nome real só usa letras/dígitos e & / - . — qualquer outro símbolo é ruído.
    if re.search(r'[^0-9A-Za-zÀ-ÿ&/.\- ]', nome):
        return ''
    # Caixa errática dentro de um token (ex.: 'SEgESS') é assinatura de OCR embaralhado.
    if re.search(r'[a-zà-ÿ][A-ZÀ-Ý]', nome):
        return ''
    if len(nome.split()) > _MAX_TOKENS_EMPRESA:
        return ''
    if not _nome_empresa_plausivel(nome):
        return ''
    return nome

def _e_continuacao_de_nome(token):
    """Fragmento em CAIXA ALTA que continua um nome quebrado em duas linhas (ex.: PIQUE)."""
    return bool(re.fullmatch(r'[A-ZÀ-ÝÇ]{3,}\.?', token))

def _montar_rotulos(linhas):
    """Recebe grupos de palavras {chave: [(left, top, texto), ...]} e devolve rótulos
    (topo, recuo_em_px, texto) ordenados de cima para baixo. O ``recuo_em_px`` deve ser
    dividido pela largura da página pelo chamador para virar fração."""
    rotulos = []
    for palavras in linhas.values():
        palavras = sorted(palavras)
        rotulos.append((min(p[1] for p in palavras), palavras[0][0], ' '.join(p[2] for p in palavras)))
    rotulos.sort()
    return rotulos

def _rotulos_numeros_texto(pagina, tolerancia_linha):
    """Extrai rótulos e números direto da camada de texto do PDF (nativo, sem OCR)."""
    from collections import defaultdict
    largura = pagina.width
    numeros = []
    linhas = defaultdict(list)
    for w in pagina.extract_words():
        texto = w['text'].strip()
        if not texto:
            continue
        if re.fullmatch(r'\d{1,5}', texto) and (w['x0'] / largura) >= _RECUO_COLUNA_QTDE_MIN:
            numeros.append((w['top'], int(texto)))
            continue
        # Agrupa por linha usando a proximidade vertical (mesma faixa de "topo").
        linhas[round(w['top'] / tolerancia_linha)].append((w['x0'], w['top'], texto))
    rotulos = [(topo, left / largura, texto) for topo, left, texto in _montar_rotulos(linhas)]
    return rotulos, numeros

def _rotulos_numeros_ocr(img_bin):
    """Extrai rótulos e números de uma página digitalizada via OCR posicional."""
    from collections import defaultdict
    largura = img_bin.shape[1]
    dados = pytesseract.image_to_data(
        img_bin, lang='por+eng', config='--psm 6 --oem 3',
        output_type=pytesseract.Output.DICT
    )
    numeros = []
    linhas = defaultdict(list)  # (bloco, par, linha) -> [(left, top, texto)]
    for i in range(len(dados['text'])):
        texto = dados['text'][i].strip()
        if not texto:
            continue
        if re.fullmatch(r'\d{1,5}', texto) and (dados['left'][i] / largura) >= _RECUO_COLUNA_QTDE_MIN:
            numeros.append((dados['top'][i], int(texto)))
            continue
        chave = (dados['block_num'][i], dados['par_num'][i], dados['line_num'][i])
        linhas[chave].append((dados['left'][i], dados['top'][i], texto))
    rotulos = [(topo, left / largura, texto) for topo, left, texto in _montar_rotulos(linhas)]
    return rotulos, numeros

def _coletar_empresas_da_pagina(rotulos, numeros, tolerancia_linha, ocorrencias):
    """Percorre os rótulos de uma página, isola os que estão na faixa de recuo das
    empresas e associa o total da mesma linha, acumulando em ``ocorrencias`` (mutado)."""
    ultimo_topo = None  # topo da última empresa registrada nesta página
    for topo, recuo, texto in rotulos:
        if not (_RECUO_EMPRESA_MIN <= recuo < _RECUO_EMPRESA_MAX):
            continue
        if _linha_de_rodape(texto):
            continue
        candidatos = sorted(
            (n for n in numeros if abs(n[0] - topo) < tolerancia_linha),
            key=lambda n: abs(n[0] - topo)
        )
        total = candidatos[0][1] if candidatos else None
        if total is None:
            # Linha na faixa de empresa mas sem número: só a tratamos como continuação
            # de um nome quebrado se for um fragmento curto em CAIXA ALTA logo abaixo da
            # empresa anterior (evita engolir rodapé/ruído).
            fragmento = texto.split()
            if (ocorrencias and ultimo_topo is not None
                    and 0 < topo - ultimo_topo < tolerancia_linha * 8
                    and len(fragmento) <= 2
                    and all(_e_continuacao_de_nome(f) for f in fragmento)):
                ocorrencias[-1][0] += ' ' + ' '.join(fragmento)
                ultimo_topo = topo
            continue
        ocorrencias.append([texto, total])
        ultimo_topo = topo

def extrair_refeicoes_por_empresa(arquivo_pdf, usar_ocr=False):
    """Extrai o total de refeições agrupado por empresa.

    Identifica as empresas pelo recuo horizontal (o nível intermediário da hierarquia
    do relatório) e associa a cada uma o total já impresso na coluna de quantidade.

    Funciona com os dois formatos do relatório:
      * PDF nativo (com camada de texto): lê as palavras e suas posições diretamente,
        sem OCR — mais rápido e sem ruído;
      * PDF digitalizado (imagem): cai para OCR posicional, detectando a rotação uma
        única vez por documento (evita páginas ilegíveis por OSD de baixa confiança).

    Marcar ``usar_ocr`` força o caminho de OCR mesmo quando há camada de texto.
    """
    from collections import OrderedDict
    ocorrencias = []  # [[texto_bruto, total], ...] na ordem em que aparecem

    with pdfplumber.open(arquivo_pdf) as pdf:
        amostra = pdf.pages[0].extract_text() or ''
        tem_camada_texto = len(amostra) > 200 and re.search(r'refei|per[ií]odo', amostra, re.IGNORECASE)

        if tem_camada_texto and not usar_ocr:
            for pagina in pdf.pages:
                tolerancia_linha = 0.012 * pagina.height
                rotulos, numeros = _rotulos_numeros_texto(pagina, tolerancia_linha)
                _coletar_empresas_da_pagina(rotulos, numeros, tolerancia_linha, ocorrencias)

        # OCR quando: sem camada de texto, forçado pelo usuário, ou a camada de texto
        # não rendeu nenhuma empresa (fallback).
        if not ocorrencias:
            rotacao = _detectar_rotacao_documento(pdf)
            total_paginas = len(pdf.pages)
            barra = st.progress(0, text="Lendo empresas do mapa de refeições...")
            for idx, pagina in enumerate(pdf.pages):
                img_bin = _preparar_imagem_ocr(pagina, rotacao=rotacao)
                tolerancia_linha = 0.012 * img_bin.shape[0]
                rotulos, numeros = _rotulos_numeros_ocr(img_bin)
                _coletar_empresas_da_pagina(rotulos, numeros, tolerancia_linha, ocorrencias)
                barra.progress((idx + 1) / total_paginas,
                               text=f"Empresas: página {idx + 1} de {total_paginas}")
            barra.empty()

    # Limpa os nomes e agrega por empresa (uma mesma empresa pode aparecer em mais de
    # uma categoria de vínculo, ex.: Matriculado e Visitante), preservando a ordem.
    totais = OrderedDict()
    for texto, total in ocorrencias:
        if total is None:
            continue
        nome = _limpar_nome_empresa(texto)
        if not nome:
            continue
        totais[nome] = totais.get(nome, 0) + total

    return [
        {"Arquivo": arquivo_pdf.name, "Empresa": nome, "Total": total}
        for nome, total in totais.items()
    ]

# ==================== EXPORTAÇÃO EXCEL ====================
def gerar_excel_formatado(df_principal, df_erros, tipo_relatorio, modo_abas):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_style = {
        "font": Font(name="Arial", size=10, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
        "alignment": Alignment(horizontal="center")
    }

    def adicionar_aba(ws, df, nome):
        ws.title = re.sub(r'[\\/*?:\[\]]', '', nome)[:31]
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = header_style["font"]
            cell.fill = header_style["fill"]
            cell.alignment = header_style["alignment"]
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            ws.append(list(row))
            for cell in ws[r_idx]:
                cell.font = Font(name="Arial", size=10)
        ws.freeze_panes = "A2"

    if modo_abas == "unica" or df_principal.empty:
        adicionar_aba(wb.create_sheet("Extração"), df_principal, "Extração Completa")
    else:
        for arquivo in df_principal['Arquivo'].unique():
            adicionar_aba(wb.create_sheet(), df_principal[df_principal['Arquivo'] == arquivo], arquivo[:30])

    if not df_erros.empty:
        ws_err = wb.create_sheet("⚠️ Linhas Rejeitadas")
        ws_err.append(["Arquivo", "Linha"])
        for cell in ws_err[1]:
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        for row in df_erros.itertuples(index=False):
            ws_err.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ==================== INTERFACE STREAMLIT ====================
st.set_page_config(page_title="Apoena Extrator", layout="wide")

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #1A2B5E 0%, #F4614D 100%);
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .stat-card {
        background: #F5F7FA;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #F4614D;
        margin: 0.5rem 0;
    }
    .success-badge {
        background: #22C55E;
        color: white;
        padding: 0.2rem 0.8rem;
        border-radius: 20px;
        display: inline-block;
    }
</style>
""", unsafe_allow_html=True)

col1, col2, col3 = st.columns([1, 5, 1])
with col2:
    st.markdown('<div class="main-header">', unsafe_allow_html=True)
    st.title("🚀 Extrator Universal de Dados - Apoena")
    st.caption("""
        <span style="color:#F4614D;">●</span> Processamento Inteligente 
        <span style="color:#0D9488;">●</span> OCR de Alta Precisão 
        <span style="color:#F59E0B;">●</span> 100% Local e Seguro
    """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 📊 Estatísticas da Sessão")
    if 'total_processados' not in st.session_state:
        st.session_state.total_processados = 0
        st.session_state.total_linhas = 0
    st.metric("📄 PDFs Processados", st.session_state.total_processados)
    st.metric("📋 Linhas Extraídas", st.session_state.total_linhas)
    st.markdown("---")

tipo = st.radio("1. Tipo de relatório:", 
                options=["fiscal", "hotel", "exames", "refeicoes"],
                format_func=lambda x: {
                    "fiscal": "Notas Fiscais (NAI 89701/92284/Misto)",
                    "hotel": "Diárias e Consumo (Plaza Hotel)",
                    "exames": "Exames Ocupacionais (Biomed)",
                    "refeicoes": "Mapa de Refeições"
                }[x])

sub_refeicoes = "geral"
if tipo == "refeicoes":
    sub_refeicoes = st.radio(
        "1b. Detalhamento do mapa de refeições:",
        options=["geral", "empresa"],
        format_func=lambda x: "Total geral (por arquivo)" if x == "geral" else "Total por empresa",
        help="'Total por empresa' lê todas as páginas e soma as refeições de cada empresa "
             "(linhas com recuo à esquerda na hierarquia do relatório)."
    )

modo_abas = st.radio("2. Organização do Excel:", ["unica", "separadas"],
                     format_func=lambda x: "Uma única aba" if x == "unica" else "Uma aba por arquivo")

usar_ocr = False
if tipo in ["fiscal", "refeicoes", "hotel"]:
    usar_ocr = st.checkbox("🔍 OCR Alta Precisão (recomendado para PDFs escaneados ou com falhas)")

arquivos = st.file_uploader("3. Selecione os PDFs:", type=['pdf'], accept_multiple_files=True)

if st.button("🚀 Extrair Dados", type="primary"):
    if not arquivos:
        st.warning("⚠️ Selecione pelo menos um arquivo PDF.")
    else:
        dados_totais = []
        rejeitados_totais = []
        stats_lista = []

        with st.spinner("Processando..."):
            for arquivo in arquivos:
                with gerenciar_memoria():
                    if tipo == "fiscal":
                        dados, rejeitados, stats = extrair_fiscal(arquivo, usar_ocr)
                        dados_totais.extend(dados)
                        rejeitados_totais.extend(rejeitados)
                        stats_lista.append(stats)
                        if stats.falhas == 0:
                            st.success(f"✅ {arquivo.name}: {stats.sucesso} linhas fiscais extraídas.")
                        else:
                            st.warning(f"⚠️ {arquivo.name}: {stats.sucesso} OK, {stats.falhas} falhas.")
                    
                    elif tipo == "hotel":
                        dados = extrair_hotel(arquivo, usar_ocr)
                        dados_totais.extend(dados)
                        if dados:
                            st.success(f"✅ {arquivo.name}: {len(dados)} itens de hotel extraídos.")
                        else:
                            st.warning(f"⚠️ {arquivo.name}: 0 itens extraídos — confira se o PDF tem consumo/diárias ou ative o OCR.")
                        stats_lista.append(EstatisticasProcessamento(arquivo=arquivo.name, sucesso=len(dados)))

                    elif tipo == "exames":
                        dados = extrair_exames(arquivo)
                        dados_totais.extend(dados)
                        if dados:
                            st.success(f"✅ {arquivo.name}: {len(dados)} exames extraídos.")
                        else:
                            st.warning(f"⚠️ {arquivo.name}: 0 exames extraídos.")
                        stats_lista.append(EstatisticasProcessamento(arquivo=arquivo.name, sucesso=len(dados)))

                    elif tipo == "refeicoes":
                        if sub_refeicoes == "empresa":
                            dados = extrair_refeicoes_por_empresa(arquivo, usar_ocr)
                            rotulo_item = "empresas"
                        else:
                            dados = extrair_refeicoes(arquivo, usar_ocr)
                            rotulo_item = "registros de refeição"
                        dados_totais.extend(dados)
                        if dados:
                            st.success(f"✅ {arquivo.name}: {len(dados)} {rotulo_item} extraídos.")
                        else:
                            st.warning(f"⚠️ {arquivo.name}: 0 {rotulo_item} extraídos — confira o PDF ou ative o OCR.")
                        stats_lista.append(EstatisticasProcessamento(arquivo=arquivo.name, sucesso=len(dados)))

        st.session_state.total_processados += len(arquivos)
        st.session_state.total_linhas += len(dados_totais)

        if stats_lista:
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📄 Arquivos processados", len(arquivos))
            with col2:
                st.metric("📋 Linhas extraídas", len(dados_totais))
            with col3:
                total_encontradas = sum(s.total_linhas_encontradas for s in stats_lista)
                total_sucesso = sum(s.sucesso for s in stats_lista)
                if total_encontradas > 0:
                    # Fiscal: usa linhas encontradas vs processadas com sucesso
                    taxa = total_sucesso / total_encontradas * 100
                else:
                    # Hotel, exames, refeições: usa arquivos com dados vs total de arquivos
                    arquivos_com_dados = sum(1 for s in stats_lista if s.sucesso > 0)
                    taxa = arquivos_com_dados / max(1, len(stats_lista)) * 100
                st.metric("✅ Taxa de sucesso", f"{taxa:.1f}%")

            # Destaca os arquivos que saíram com zero registros (culpados pela taxa < 100%)
            arquivos_zerados = [s.arquivo for s in stats_lista if s.sucesso == 0]
            if arquivos_zerados:
                st.error(
                    f"❌ {len(arquivos_zerados)} arquivo(s) sem nenhum dado extraído:\n\n"
                    + "\n".join(f"- {nome}" for nome in arquivos_zerados)
                    + "\n\nAbra esses PDFs para conferir se realmente têm dados; se tiverem, "
                      "tente marcar o OCR de Alta Precisão."
                )

        # Quando o mapa de refeições é detalhado por empresa, o layout de saída muda.
        tipo_saida = "refeicoes_empresas" if (tipo == "refeicoes" and sub_refeicoes == "empresa") else tipo

        if dados_totais:
            df = pd.DataFrame(dados_totais, columns=COLUNAS_CONFIG[tipo_saida])
            df_erros = pd.DataFrame(rejeitados_totais) if rejeitados_totais else pd.DataFrame()
            excel_buffer = gerar_excel_formatado(df, df_erros, tipo_saida, modo_abas)
            with st.expander(f"👁 Prévia dos dados ({min(5, len(df))} primeiras linhas)", expanded=True):
                st.dataframe(df.head(5), use_container_width=True)
            st.download_button(
                label="📥 Baixar Excel",
                data=excel_buffer,
                file_name=f"Extracao_{tipo_saida}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("ℹ️ Nenhum dado foi encontrado nos arquivos enviados para o tipo selecionado.")
